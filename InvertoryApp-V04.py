import openpyxl
import tkinter as tk
from tkinter import messagebox, filedialog, ttk
from openpyxl import Workbook
import customtkinter as ctk

class StoreManagementApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Store Management App")
        self.root.geometry("1000x800")
        self.product_list = []
        self.item_no_counter = 1
        self.requisition_item_counter=1
        self.setup_frames()
        self.setup_home_frame()
        self.setup_add_product_frame()
        self.setup_withdraw_product_frame()
        self.show_home_frame()
        self.load_from_excel()

    def setup_frames(self):
        self.frame_home = ctk.CTkFrame(self.root)
        self.frame_add_product = ctk.CTkFrame(self.root)
        self.frame_withdraw_product = ctk.CTkFrame(self.root)

    def setup_home_frame(self):
        self.frame_home.pack(pady=10, padx=10, fill="both", expand=True)
        
        ctk.CTkLabel(self.frame_home, text="Super Petrochemical Limited", font=("Arial", 24, "bold")).pack(pady=10)
        ctk.CTkLabel(self.frame_home, text="Instrument and Control Department", font=("Arial", 15, "bold")).pack(pady=10)
        
        ctk.CTkButton(self.frame_home, text="Add Product", command=self.show_add_product_frame).pack(pady=10)
        ctk.CTkButton(self.frame_home, text="Withdraw Product", command=self.show_withdraw_product_frame).pack(pady=10)

    def setup_add_product_frame(self):
        self.frame_logo_add = ctk.CTkFrame(self.frame_add_product)
        self.frame_logo_add.pack(pady=10, padx=10, fill="x")

        self.company_name_add = ctk.CTkLabel(self.frame_logo_add, text="Super Petrochemical Limited", font=("Arial", 24, "bold"))
        self.company_name_add.pack(pady=5)
        self.department_name_add = ctk.CTkLabel(self.frame_logo_add, text="Instrument and Control Department", font=("Arial", 15, "bold"))
        self.department_name_add.pack(pady=5)

        self.frame_top_add = ctk.CTkFrame(self.frame_add_product)
        self.frame_top_add.pack(pady=10, padx=10, fill="x")
        
        self.frame_middle_add = ctk.CTkFrame(self.frame_add_product)
        self.frame_middle_add.pack(pady=10, padx=10, fill="x")
        
        self.frame_bottom_add = ctk.CTkFrame(self.frame_add_product)
        self.frame_bottom_add.pack(pady=10, padx=10, fill="both", expand=True)

        # Top Frame Widgets
        self.frame_entries_add = ctk.CTkFrame(self.frame_top_add)
        self.frame_entries_add.pack(pady=10, padx=10)

        ctk.CTkLabel(self.frame_entries_add, text="Product Name:").grid(row=0, column=0, padx=5, pady=5)
        self.entry_product_name_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_product_name_add.grid(row=0, column=1, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_add, text="Tag:").grid(row=0, column=2, padx=5, pady=5)
        self.entry_tag_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_tag_add.grid(row=0, column=3, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_add, text="Plant Name:").grid(row=1, column=0, padx=5, pady=5)
        self.entry_plant_name_add = ctk.CTkComboBox(self.frame_entries_add, values=["CHEMAX", "ZEHUA", "BTX", "HEXANE", "LCP", "TNS", "REFORMER 2", "CFU-10000BPD"])
        self.entry_plant_name_add.grid(row=1, column=1, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_add, text="Store:").grid(row=1, column=2, padx=5, pady=5)
        self.entry_store_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_store_add.grid(row=1, column=3, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_add, text="Type:").grid(row=2, column=0, padx=5, pady=5)
        self.entry_type_add = ctk.CTkComboBox(self.frame_entries_add, values=["PG", "PT", "LT","LG","FT","FI","TG","TT","TG","TE","SENSOR","PLC","IO CARD","PSU","SWITCH","IO CARD","INDICATOR","BUTTON",])
        self.entry_type_add.grid(row=2, column=1, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_add, text="Size:").grid(row=2, column=2, padx=5, pady=5)
        self.entry_size_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_size_add.grid(row=2, column=3, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_add, text="Range:").grid(row=3, column=0, padx=5, pady=5)
        self.entry_range_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_range_add.grid(row=3, column=1, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_add, text="Quantity:").grid(row=3, column=2, padx=5, pady=5)
        self.entry_quantity_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_quantity_add.grid(row=3, column=3, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_add, text="Rack:").grid(row=4, column=0, padx=5, pady=5)
        self.entry_rack_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_rack_add.grid(row=4, column=1, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_add, text="Column:").grid(row=4, column=2, padx=5, pady=5)
        self.entry_column_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_column_add.grid(row=4, column=3, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_add, text="Box No:").grid(row=5, column=0, padx=5, pady=5)
        self.entry_box_no_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_box_no_add.grid(row=5, column=1, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_add, text="Part No:").grid(row=5, column=2, padx=5, pady=5)
        self.entry_part_no_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_part_no_add.grid(row=5, column=3, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_add, text="Description:").grid(row=6, column=0, padx=5, pady=5)
        self.entry_description_add = tk.Text(self.frame_entries_add, height=5, width=40)
        self.entry_description_add.grid(row=6, column=1, columnspan=3, padx=5, pady=5)

        self.button_add_product = ctk.CTkButton(self.frame_entries_add, text="Add Product", command=self.add_product)
        self.button_add_product.grid(row=7, column=0, columnspan=4, pady=10)

        # Middle Frame Widgets
        ctk.CTkLabel(self.frame_middle_add, text="Search Product:").grid(row=0, column=0, padx=5, pady=5)
        self.entry_search_add = ctk.CTkEntry(self.frame_middle_add)
        self.entry_search_add.grid(row=0, column=1, padx=5, pady=5)
        
        ctk.CTkButton(self.frame_middle_add, text="Search", command=self.search_product).grid(row=0, column=2, padx=5, pady=5)
        ctk.CTkButton(self.frame_middle_add, text="Clear Search", command=self.refresh_product_list).grid(row=0, column=3, padx=5, pady=5)

        # Bottom Frame Widgets
        self.tree_add = ttk.Treeview(self.frame_bottom_add, columns=(
            "Item No", "Name","Quantity","Size","Type","Range","Store","Rack","Column","Box", "Description", "Plant Name", "Part No", "Tag"), show="headings")
        self.tree_add.heading("Item No", text="Item No")
        self.tree_add.heading("Name", text="Product Name")
        self.tree_add.heading("Quantity", text="Quantity")
        self.tree_add.heading("Size", text="Size")
        self.tree_add.heading("Type", text="Type")
        self.tree_add.heading("Range", text="Range")
        self.tree_add.heading("Store", text="Store")
        self.tree_add.heading("Rack", text="Rack")
        self.tree_add.heading("Column", text="Column")
        self.tree_add.heading("Box", text="Box")
        self.tree_add.heading("Description", text="Description")
        self.tree_add.heading("Plant Name", text="Plant Name")
        self.tree_add.heading("Part No", text="Part No")
        self.tree_add.heading("Tag", text="Tag")

        self.tree_add.pack(fill="both", expand=True)
        ctk.CTkButton(self.frame_bottom_add, text="Back to Home", command=self.show_home_frame).pack(pady=10)
        ctk.CTkButton(self.frame_bottom_add, text="Print", command=self.print_popup).pack(pady=10)
        self.tree_add.bind("<Double-1>", self.load_selected_product_add)
        self.refresh_product_list()

    def setup_withdraw_product_frame(self):
        self.frame_logo_withdraw = ctk.CTkFrame(self.frame_withdraw_product)
        self.frame_logo_withdraw.pack(pady=10, padx=10, fill="x")

        self.company_name_withdraw = ctk.CTkLabel(self.frame_logo_withdraw, text="Super Petrochemical Limited", font=("Arial", 24, "bold"))
        self.company_name_withdraw.pack(pady=5)
        self.department_name_withdraw = ctk.CTkLabel(self.frame_logo_withdraw, text="Instrument and Control Department", font=("Arial", 15, "bold"))
        self.department_name_withdraw.pack(pady=5)

        self.frame_top_withdraw = ctk.CTkFrame(self.frame_withdraw_product)
        self.frame_top_withdraw.pack(pady=10, padx=10, fill="x")
        
        self.frame_middle_withdraw = ctk.CTkFrame(self.frame_withdraw_product)
        self.frame_middle_withdraw.pack(pady=10, padx=10, fill="x")
        
        self.frame_bottom_withdraw = ctk.CTkFrame(self.frame_withdraw_product)
        self.frame_bottom_withdraw.pack(pady=10, padx=10, fill="both", expand=True)

        # Top Frame Widgets
        self.frame_entries_withdraw = ctk.CTkFrame(self.frame_top_withdraw)
        self.frame_entries_withdraw.pack(pady=10, padx=10)

        ctk.CTkLabel(self.frame_entries_withdraw, text="SR By:").grid(row=0, column=0, padx=5, pady=5)
        self.sr_by_requisition = ctk.CTkEntry(self.frame_entries_withdraw)
        self.sr_by_requisition.grid(row=0, column=1, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_withdraw, text="Required for:").grid(row=0, column=2, padx=5, pady=5)
        self.required_requisition = ctk.CTkEntry(self.frame_entries_withdraw)
        self.required_requisition.grid(row=0, column=3, padx=5, pady=5)
        # Requisited Items Table
        self.tree_requisition = ttk.Treeview(self.frame_top_withdraw, columns=("SL No.", "Item No","Description","Quantity","Size","Remarks"), show="headings")
        self.tree_requisition.column('SL No.',width=30)
        self.tree_requisition.column('Item No',width=100)
        self.tree_requisition.column('Description',width=400)
        self.tree_requisition.column('Quantity',width=100)
        self.tree_requisition.column('Size',width=100)
        self.tree_requisition.column('Remarks',width=100)
        self.tree_requisition.heading("SL No.", text="Sl No.")
        self.tree_requisition.heading("Item No", text="Item No")
        self.tree_requisition.heading("Description", text="Description")
        self.tree_requisition.heading("Quantity", text="Quantity")
        self.tree_requisition.heading("Size", text="Size")
        self.tree_requisition.heading("Remarks", text="Remarks")
        self.tree_requisition.pack(fill="x", expand=True)

        
        # ctk.CTkLabel(self.frame_entries_withdraw, text="Quantity:").grid(row=1, column=0, padx=5, pady=5)
        # self.entry_quantity_withdraw = ctk.CTkEntry(self.frame_entries_withdraw)
        # self.entry_quantity_withdraw.grid(row=1, column=1, padx=5, pady=5)

        self.button_withdraw_product = ctk.CTkButton(self.frame_top_withdraw, text="Withdraw Product", command=self.withdraw_product)
        self.button_withdraw_product.pack( padx=10, pady=10)

        # Middle Frame Widgets
        ctk.CTkLabel(self.frame_middle_withdraw, text="Search Product:").grid(row=0, column=0, padx=5, pady=5)
        self.entry_search_withdraw = ctk.CTkEntry(self.frame_middle_withdraw)
        self.entry_search_withdraw.grid(row=0, column=1, padx=5, pady=5)
        
        ctk.CTkButton(self.frame_middle_withdraw, text="Search", command=self.search_product).grid(row=0, column=2, padx=5, pady=5)
        ctk.CTkButton(self.frame_middle_withdraw, text="Clear Search", command=self.refresh_product_list_withdraw).grid(row=0, column=3, padx=5, pady=5)
        ctk.CTkLabel(self.frame_bottom_withdraw, text="Stock Items").pack(padx=10, pady = 10, fill="x")
        self.tree_withdraw = ttk.Treeview(self.frame_bottom_withdraw, columns=(
            "Item No", "Name","Quantity","Size","Type","Range","Store","Rack","Column","Box", "Description", "Plant Name", "Part No", "Tag"), show="headings")
        self.tree_withdraw.heading("Item No", text="Item No")
        self.tree_withdraw.heading("Name", text="Product Name")
        self.tree_withdraw.heading("Quantity", text="Quantity")
        self.tree_withdraw.heading("Size", text="Size")
        self.tree_withdraw.heading("Type", text="Type")
        self.tree_withdraw.heading("Range", text="Range")
        self.tree_withdraw.heading("Store", text="Store")
        self.tree_withdraw.heading("Rack", text="Rack")
        self.tree_withdraw.heading("Column", text="Column")
        self.tree_withdraw.heading("Box", text="Box")
        self.tree_withdraw.heading("Description", text="Description")
        self.tree_withdraw.heading("Plant Name", text="Plant Name")
        self.tree_withdraw.heading("Part No", text="Part No")
        self.tree_withdraw.heading("Tag", text="Tag")
        self.tree_withdraw.pack(fill="both", expand=True)
        # self.tree_withdraw.insert("", "end",("Item No","1"))
        ctk.CTkButton(self.frame_withdraw_product, text="Back to Home", command=self.show_home_frame).pack(pady=10)
        ctk.CTkButton(self.frame_withdraw_product, text="Print", command=self.print_popup).pack(pady=10)
        self.tree_withdraw.bind("<Double-1>", lambda event: self.add_to_cart())
        
        self.refresh_product_list_withdraw()
        

    def show_home_frame(self):
        self.frame_home.pack(pady=10, padx=10, fill="both", expand=True)
        self.frame_add_product.pack_forget()
        self.frame_withdraw_product.pack_forget()

    def show_add_product_frame(self):
        self.frame_add_product.pack(pady=10, padx=10, fill="both", expand=True)
        self.frame_home.pack_forget()
        self.frame_withdraw_product.pack_forget()

    def show_withdraw_product_frame(self):
        self.frame_withdraw_product.pack(pady=10, padx=10, fill="both", expand=True)
        self.frame_home.pack_forget()
        self.frame_add_product.pack_forget()
    def export_to_excel(self):
        file_path = "products.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Products"
        headers = ["Item No", "Product Name", "Quantity", "Size", "Type", "Range", "Store", "Rack", "Column", "Box", "Description", "Plant Name", "Part No", "Tag"]
        sheet.append(headers)

        for product in self.product_list:
            row = [product[header] for header in headers]
            sheet.append(row)

        try:
            workbook.save(file_path)
            # messagebox.showinfo("Export Successful", f"Data exported successfully to {file_path}")
        except Exception as e:
            messagebox.showerror("Export Error", f"An error occurred while exporting: {e}")
        
        workbook.save("products.xlsx")
        messagebox.showinfo("Excel File", "Products saved to products.xlsx")

    def add_product(self):
        product_name = self.entry_product_name_add.get()
        tag = self.entry_tag_add.get()
        plant_name = self.entry_plant_name_add.get()
        store = self.entry_store_add.get()
        product_type = self.entry_type_add.get()
        size = self.entry_size_add.get()
        product_range = self.entry_range_add.get()
        quantity = self.entry_quantity_add.get()
        rack = self.entry_rack_add.get()
        column = self.entry_column_add.get()
        box_no = self.entry_box_no_add.get()
        description = self.entry_description_add.get("1.0", tk.END).strip()
        # part_no = self.entry_part_no_add.get()
        part_no = self.generate_part_no(plant_name, product_type, product_range)

        if not all([product_name, tag, plant_name, store, product_type, size, product_range, quantity, rack, column, box_no, description, part_no]):
            messagebox.showerror("Error", "Please fill all the fields.")
            return

        item_no = self.item_no_counter
        self.item_no_counter += 1

        self.product_list.append({
            "Item No": item_no,
            "Product Name": product_name,
            "Quantity": quantity,
            "Size": size,
            "Type": product_type,
            "Range": product_range,
            "Store": store,
            "Rack": rack,
            "Column": column,
            "Box": box_no,
            "Description": description,
            "Plant Name": plant_name,
            "Part No": part_no,
            "Tag": tag
        })
        self.export_to_excel()
        self.refresh_product_list()
        
        # messagebox.showinfo("Success", "Product added successfully.")
        self.clear_add_product_entries()

    def withdraw_product(self):
        pass
    #     product_name = self.entry_product_name_withdraw.get()
    #     tag = self.entry_tag_withdraw.get()
    #     quantity = self.entry_quantity_withdraw.get()

    #     if not all([product_name, tag, quantity]):
    #         messagebox.showerror("Error", "Please fill all the fields.")
    #         return

    #     quantity = int(quantity)
    #     product_found = False

    #     for product in self.product_list:
    #         if product["Product Name"] == product_name and product["Tag"] == tag:
    #             product_found = True
    #             if product["Quantity"] >= quantity:
    #                 product["Quantity"] -= quantity
    #                 self.refresh_product_list()
    #                 messagebox.showinfo("Success", "Product withdrawn successfully.")
    #             else:
    #                 messagebox.showerror("Error", "Insufficient quantity in stock.")
    #             break

    #     if not product_found:
    #         messagebox.showerror("Error", "Product not found.")
    def load_from_excel(self):
            file_path = "products.xlsx"
            if not file_path:
                return
            
            try:
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active

                self.product_list = []
                self.item_no_counter = 1

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    product = {
                        "Item No": row[0],
                        "Product Name": row[1],
                        "Quantity": row[2],
                        "Size": row[3],
                        "Type": row[4],
                        "Range": row[5],
                        "Store": row[6],
                        "Rack": row[7],
                        "Column": row[8],
                        "Box": row[9],
                        "Description": row[10],
                        "Plant Name": row[11],
                        "Part No": row[12],
                        "Tag": row[13]
                    }
                    self.product_list.append(product)
                    self.item_no_counter += 1

                self.refresh_product_list()
                # messagebox.showinfo("Load Successful", f"Data loaded successfully from {file_path}")
            except Exception as e:
                messagebox.showerror("Load Error", f"An error occurred while loading: {e}")

    def search_product(self):
            # search_term = self.entry_search_add.get() if isinstance(self.entry_search_add.get(), str) else str(self.entry_search_add.get()).lower()
            search_term = self.entry_search_withdraw.get() if isinstance(self.entry_search_withdraw.get(), str) else str(self.entry_search_withdraw.get()).lower()
            

            for row in self.tree_add.get_children():
                self.tree_add.delete(row)
            for row in self.tree_withdraw.get_children():
                self.tree_withdraw.delete(row)

            for product in self.product_list:
                if (
                    search_term in str(product["Item No"]) or
                    search_term in product["Product Name"].lower() or
                    search_term in product["Plant Name"].lower() or
                    search_term in product["Type"].lower() or
                    search_term in product["Description"].lower() or
                    search_term in product["Tag"].lower() or
                    search_term in product["Part No"].lower()):
                    self.tree_add.insert("", "end", values=(
                        product["Item No"], 
                        product["Product Name"], 
                        product["Quantity"],
                        product["Size"], 
                        product["Type"], 
                        product["Range"], 
                        product["Store"], 
                        product["Rack"], 
                        product["Column"], 
                        product["Box"], 
                        product["Description"], 
                        product["Plant Name"], 
                        product["Part No"], 
                        product["Tag"]))
                    self.tree_withdraw.insert("", "end", values=(
                        product["Item No"], 
                        product["Product Name"], 
                        product["Quantity"],
                        product["Size"], 
                        product["Type"], 
                        product["Range"], 
                        product["Store"], 
                        product["Rack"], 
                        product["Column"], 
                        product["Box"], 
                        product["Description"], 
                        product["Plant Name"], 
                        product["Part No"], 
                        product["Tag"]))
            
    # def search_product(self):
    #     search_term =[]
    #     if type(search_term) is int:
    #         try:
    #             search_term = self.entry_search_add.get()
    #             search_term = self.entry_search_withdraw.get()
    #         except:
    #             search_term = self.entry_search_add.get().lower()
    #             search_term = self.entry_search_withdraw.get().lower()

    #     for row in self.tree_add.get_children():
    #         self.tree_add.delete(row)
    #     for row in self.tree_withdraw.get_children():
    #         self.tree_withdraw.delete(row)

    #     for product in self.product_list:
    #         if (
    #                 search_term in product["Item No"] or
    #                 search_term in product["Tag"].lower() or
    #                 search_term in product["Product Name"].lower() or
    #                 search_term in product["Size"].lower() or
    #                 search_term in product["Plant Name"].lower() or
    #                 search_term in product["Type"].lower() or
    #                 search_term in product["Range"].lower() or
    #                 search_term in product["Store"].lower() or
    #                 search_term in product["Rack"].lower() or
    #                 search_term in product["Column"].lower() or
    #                 search_term in product["Box"].lower() or
    #                 search_term in product["Description"].lower() or
                    # search_term in product["Part No"].lower()):
                    # self.tree_add.insert("", tk.END, values=product)
                    # self.tree_add.insert("", "end", values=product)
                    # self.tree_withdraw.insert("", "end", values=product)
    def refresh_product_list(self):
        for i in self.tree_add.get_children():
            self.tree_add.delete(i)
        for product in self.product_list:
            self.tree_add.insert("", "end", values=(
                product["Item No"], product["Product Name"], product["Quantity"],product["Size"], product["Type"], product["Range"], product["Store"], product["Rack"], product["Column"], product["Box"], product["Description"], product["Plant Name"], product["Part No"], product["Tag"]
            ))
    def refresh_product_list_withdraw(self):
        self.load_from_excel()
        for i in self.tree_withdraw.get_children():
            self.tree_withdraw.delete(i)
        # print(f"Product List:{self.product_list}")
        for product in self.product_list:
            self.tree_withdraw.insert("", "end", values=(
                    product["Item No"], product["Product Name"], product["Quantity"],product["Size"], product["Type"], product["Range"], product["Store"], product["Rack"], product["Column"], product["Box"], product["Description"], product["Plant Name"], product["Part No"], product["Tag"]
                ))

       
    def load_selected_product_add(self, event):
        selected_item = self.tree_add.selection()
        if not selected_item:
            return
        
        item_index = self.tree_add.index(selected_item)
        selected_product = self.product_list[item_index]
        self.entry_product_name_add.delete(0, tk.END)
        self.entry_product_name_add.insert(0, selected_product["Product Name"])
        self.entry_quantity_add.delete(0, tk.END)
        self.entry_quantity_add.insert(0, selected_product["Quantity"])
        self.entry_type_add.set(selected_product["Type"])
        self.entry_range_add.delete(0, tk.END)
        self.entry_range_add.insert(0, selected_product["Range"])
        self.entry_store_add.delete(0, tk.END)
        self.entry_store_add.insert(0, selected_product["Store"])
        self.entry_rack_add.delete(0, tk.END)
        self.entry_rack_add.insert(0, selected_product["Rack"])
        self.entry_column_add.delete(0, tk.END)
        self.entry_column_add.insert(0, selected_product["Column"])
        self.entry_box_no_add.delete(0, tk.END)
        self.entry_box_no_add.insert(0, selected_product["Box"])
        self.entry_plant_name_add.set(selected_product["Plant Name"])
        self.entry_part_no_add.delete(0, tk.END)
        self.entry_part_no_add.insert(0, selected_product["Part No"])
        self.entry_tag_add.delete(0, tk.END)
        self.entry_tag_add.insert(0, selected_product["Tag"])
        self.entry_size_add.delete(0, tk.END)
        self.entry_size_add.insert(0, selected_product["Size"])
        self.entry_description_add.delete('1.0', 'end')
        self.entry_description_add.insert("1.0", selected_product["Description"])
    def clear_add_product_entries(self):
        self.entry_product_name_add.delete(0, 'end')
        self.entry_tag_add.delete(0, 'end')
        self.entry_plant_name_add.set("")
        self.entry_store_add.delete(0, 'end')
        self.entry_type_add.set("")
        self.entry_size_add.delete(0, 'end')
        self.entry_range_add.delete(0, 'end')
        self.entry_quantity_add.delete(0, 'end')
        self.entry_rack_add.delete(0, 'end')
        self.entry_column_add.delete(0, 'end')
        self.entry_description_add.delete('1.0', 'end')
        self.entry_part_no_add.delete(0, 'end')
        self.entry_box_no_add.delete(0, 'end')
        self.refresh_product_list()
    
    def generate_part_no(self, plant_name, type_, range_):
        plant_mapping = {
            "CHEMAX": "CHE",
            "ZEHUA": "ZEH",
            "BTX": "BTX",
            "HEXANE": "HEX",
            "LCP": "LCP",
            "TNS": "TNS",
            "REFORMER 2": "REF02",
            "CFU-10000BPD": "CFU"
        }
        plant_code = plant_mapping.get(plant_name.upper(), plant_name[:3].upper())
        part_no = f"{plant_code}-{type_.upper()}-{range_}"
        return part_no
    def add_to_cart(self):
        selected_item = self.tree_withdraw.selection()
        if not selected_item:
            return
         
        curntItem = self.tree_withdraw.focus()

        item_index = self.tree_withdraw.index(curntItem)
        curntItemNo = self.tree_withdraw.item(curntItem)["values"][0]
        print(f"Current Item No. {curntItemNo}")
        i=0
        for product in self.product_list:
                    if (
                    str(curntItemNo) in str(product["Item No"])):
                        print(product)
                        
        # selected_product = self.product_list[curntItemIndex-1]
        # if(self.tree_requisition.get_children() == "NULL"):
        #     self.requisition_item_counter=1
        # self.tree_requisition.insert("", "end", values=(self.requisition_item_counter,selected_product["Item No"], selected_product["Product Name"], selected_product["Quantity"],selected_product["Size"]))
        # self.requisition_item_counter +=1

        
        
    def print_popup(self):
        popup = tk.Toplevel(self.root)
        popup.title("Print")
        popup.geometry("300x200")

        def print_details():
            # Code to handle printing can go here
            print("Print initiated")
            popup.destroy()

        tk.Label(popup, text="Select print options:", font=("Arial", 14)).pack(pady=10)

        # Add print options here (e.g., selection of specific details, number of copies)
        tk.Label(popup, text="Number of copies:").pack(pady=5)
        num_copies = tk.Entry(popup)
        num_copies.pack(pady=5)

        tk.Button(popup, text="Print", command=print_details).pack(pady=10)
        tk.Button(popup, text="Cancel", command=popup.destroy).pack(pady=10)
    
        
if __name__ == "__main__":
    ctk.set_appearance_mode("System")  # Modes: "System" (default), "Dark", "Light"
    ctk.set_default_color_theme("green")  # Themes: "blue" (default), "green", "dark-blue"
    root = ctk.CTk()
    app = StoreManagementApp(root)
    root.mainloop()