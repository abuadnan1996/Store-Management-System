import openpyxl
import tkinter as tk
from tkinter import messagebox, filedialog, ttk
from openpyxl import Workbook
import customtkinter as ctk
import mysql.connector
from mysql.connector import Error

class StoreManagementApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Store Management App")
        self.root.geometry("1040x800")
        self.product_list = []
        self.previous_product_list = []
        self.cart_list = []
        self.numbers =[]
        self.available_numbers = []
        self.product_list_db=[]
        self.product_list_test=[]
        self.fields = ("Item No", "Product Name", "Brand", "Model", "Quantity", "Size", "Type", "Range", "Store", "Rack", 
                         "Column", "Row", "Box", "Description", "Plant Name", "Part No", "Tag")
        self.db_fields = ("product_name", "brand", "model", "quantity", "size", "type", "range", "store", "rack", 
                         "column", "row", "box", "description", "plant", "part_no", "tag")
        self.popup_window=0
        self.item_no_counter = 1
        self.requisition_item_counter=1
        self.product_loaded_add=0
        #Boolean Flag for Withdraw window if searched or not
        self.searched_withdraw=0
        self.searched_add=0
        self.setup_frames()
        self.setup_home_frame()
        self.setup_add_product_frame()
        self.setup_withdraw_product_frame()
        self.show_home_frame()
        self.create_db_connection("localhost","root","Adnan1996","instrument_store")
        self.load_from_excel()
        self.load_from_db('instrument_store.stocks')
        
        # print(self.product_list)
        for products in self.product_list:
            values = tuple(products.values())
            self.product_list_db.append(values)
        # print(f"Products in tuples::: {self.product_list_db}")
        # self.add_to_db("stocks",self.product_list_db)
        
    def setup_frames(self):
        self.frame_home = ctk.CTkFrame(self.root)
        self.frame_add_product = ctk.CTkFrame(self.root)
        self.frame_withdraw_product = ctk.CTkFrame(self.root)

    def setup_home_frame(self):
        self.frame_home.pack(pady=10, padx=10, fill="both", expand=True)
        
        ctk.CTkLabel(self.frame_home, text="Super Petrochemical Limited", font=("Arial", 24, "bold")).pack(pady=10)
        ctk.CTkLabel(self.frame_home, text="Instrument and Control Department", font=("Arial", 15, "bold")).pack(pady=10)
        
        ctk.CTkButton(self.frame_home, text="Add Product", command=self.show_add_product_frame).pack(pady=10)
        ctk.CTkButton(self.frame_home, text="Withdraw Product", command=self.show_withdraw_product_frame).pack(pady=10)

    def setup_add_product_frame(self):
        self.top_button_add = ctk.CTkFrame(self.frame_add_product)
        self.top_button_add.pack(padx=10)
        ctk.CTkButton(self.top_button_add, text="Back to Home", command=self.show_home_frame).pack()

        self.frame_logo_add = ctk.CTkFrame(self.frame_add_product)
        self.frame_logo_add.pack(pady=10, padx=10, fill="x")
        

        self.company_name_add = ctk.CTkLabel(self.frame_logo_add, text="Super Petrochemical Limited", font=("Arial", 24, "bold"))
        self.company_name_add.pack(padx=10)

        self.department_name_add = ctk.CTkLabel(self.frame_logo_add, text="Instrument and Control Department", font=("Arial", 15, "bold"))
        self.department_name_add.pack(padx=10)

        self.frame_top_add = ctk.CTkFrame(self.frame_add_product)
        self.frame_top_add.pack(pady=10, padx=10, fill="x")
        
        self.frame_middle_add = ctk.CTkFrame(self.frame_add_product)
        self.frame_middle_add.pack(pady=10, padx=10, fill="x")
        
        self.frame_bottom_add = ctk.CTkFrame(self.frame_add_product)
        self.frame_bottom_add.pack(pady=10, padx=10, fill="both", expand=True)

        # Top Frame Widgets
        self.frame_entries_add = ctk.CTkFrame(self.frame_top_add)
        self.frame_entries_add.pack(pady=10, padx=10)

         # Row 0
        ctk.CTkLabel(self.frame_entries_add, text="Product Name:").grid(row=0, column=0, padx=5, pady=5)
        self.entry_product_name_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_product_name_add.grid(row=0, column=1, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_add, text="Brand:").grid(row=0, column=2, padx=5, pady=5)
        self.entry_brand_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_brand_add.grid(row=0, column=3, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_add, text="Model:").grid(row=0, column=4, padx=5, pady=5)
        self.entry_model_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_model_add.grid(row=0, column=5, padx=5, pady=5)
        # Row 1
        ctk.CTkLabel(self.frame_entries_add, text="Quantity:").grid(row=1, column=0, padx=5, pady=5)
        self.entry_quantity_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_quantity_add.grid(row=1, column=1, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_add, text="Size:").grid(row=1, column=2, padx=5, pady=5)
        self.entry_size_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_size_add.grid(row=1, column=3, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_add, text="Type:").grid(row=1, column=4, padx=5, pady=5)
        self.entry_type_add = ctk.CTkComboBox(self.frame_entries_add, values=["PG", "PT", "LT", "LG", "FT", "FI", "TG", "TT", "TG", "TE", "SENSOR", "PLC", "IO CARD", "PSU", "SWITCH", "INDICATOR", "BUTTON"])
        self.entry_type_add.grid(row=1, column=5, padx=5, pady=5)


        # Row 2
        ctk.CTkLabel(self.frame_entries_add, text="Range:").grid(row=2, column=0, padx=5, pady=5)
        self.entry_range_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_range_add.grid(row=2, column=1, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_add, text="Store:").grid(row=2, column=2, padx=5, pady=5)
        self.entry_store_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_store_add.grid(row=2, column=3, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_add, text="Rack:").grid(row=2, column=4, padx=5, pady=5)
        self.entry_rack_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_rack_add.grid(row=2, column=5, padx=5, pady=5)

        # Row 3
        ctk.CTkLabel(self.frame_entries_add, text="Row:").grid(row=3, column=0, padx=5, pady=5)
        self.entry_row_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_row_add.grid(row=3, column=1, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_add, text="Column:").grid(row=3, column=2, padx=5, pady=5)
        self.entry_column_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_column_add.grid(row=3, column=3, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_add, text="Box No:").grid(row=3, column=4, padx=5, pady=5)
        self.entry_box_no_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_box_no_add.grid(row=3, column=5, padx=5, pady=5)

        # Row 4
        ctk.CTkLabel(self.frame_entries_add, text="Plant Name:").grid(row=4, column=0, padx=5, pady=5)
        self.entry_plant_name_add = ctk.CTkComboBox(self.frame_entries_add, values=["CHEMEX", "ZEHUA", "BTX", "HEXANE", "LCP", "TNS", "REFORMER 2", "CFU-10000BPD"])
        self.entry_plant_name_add.grid(row=4, column=1, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_add, text="Part No.:").grid(row=4, column=2, padx=5, pady=5)
        self.entry_part_no_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_part_no_add.grid(row=4, column=3, padx=5, pady=5)
        
        ctk.CTkLabel(self.frame_entries_add, text="Tag:").grid(row=4, column=4, padx=5, pady=5)
        self.entry_tag_add = ctk.CTkEntry(self.frame_entries_add)
        self.entry_tag_add.grid(row=4, column=5, padx=5, pady=5)

        # Row 8
        ctk.CTkLabel(self.frame_entries_add, text="Description:").grid(row=5, column=0, padx=5, pady=5)
        self.entry_description_add = tk.Text(self.frame_entries_add, height=5, width=40)
        self.entry_description_add.grid(row=5, column=1, columnspan=6, padx=5, pady=5)

        self.button_add_product = ctk.CTkButton(self.frame_entries_add, text="Add Product", command=self.add_product)
        self.button_add_product.grid(row=6, column=0, columnspan=6, pady=10)
        self.button_clear_entry = ctk.CTkButton(self.frame_entries_add, text="Clear", command=self.clear_add_product_entries)
        self.button_clear_entry.grid(row=7, column=0, columnspan=6, pady=10)

        # Middle Frame Widgets
        ctk.CTkLabel(self.frame_middle_add, text="Search Product:").grid(row=0, column=0, padx=5, pady=5)
        self.entry_search_add = ctk.CTkEntry(self.frame_middle_add)
        self.entry_search_add.grid(row=0, column=1, padx=5, pady=5)
        
        ctk.CTkButton(self.frame_middle_add, text="Search", command=self.search_product_add).grid(row=0, column=2, padx=5, pady=5)
        ctk.CTkButton(self.frame_middle_add, text="Clear", command=self.refresh_product_list).grid(row=0, column=3, padx=5, pady=5)

        # Bottom Frame Widgets
        self.tree_add = ttk.Treeview(self.frame_bottom_add, columns=self.fields, show="headings")

        
        verscrlbar = ttk.Scrollbar(self.frame_bottom_add, 
                           orient ="vertical", 
                           command = self.tree_add.yview)
        # Calling pack method w.r.to vertical 
        # scrollbar
        verscrlbar.pack(side ='right', fill ='both')
        #resizing the scrollbar according to table size
        self.tree_add.configure(yscrollcommand=verscrlbar.set)

        self.tree_add.heading("Item No", text="Item No")
        self.tree_add.heading("Product Name", text="Product Name")
        self.tree_add.heading("Brand", text="Brand")
        self.tree_add.heading("Model", text="Model")
        self.tree_add.heading("Quantity", text="Quantity")
        self.tree_add.heading("Size", text="Size")
        self.tree_add.heading("Type", text="Type")
        self.tree_add.heading("Range", text="Range")
        self.tree_add.heading("Store", text="Store")
        self.tree_add.heading("Rack", text="Rack")
        self.tree_add.heading("Column", text="Column")
        self.tree_add.heading("Row", text="Row")
        self.tree_add.heading("Box", text="Box")
        self.tree_add.heading("Description", text="Description")
        self.tree_add.heading("Plant Name", text="Plant Name")
        self.tree_add.heading("Part No", text="Part No")
        self.tree_add.heading("Tag", text="Tag")


        self.tree_add.pack(fill="both", expand=True)
        # ctk.CTkButton(self.frame_bottom_add, text="Back to Home", command=self.show_home_frame).pack(pady=10)
        # ctk.CTkButton(self.frame_bottom_add, text="Print", command=self.print_popup).pack(pady=10)
        self.tree_add.bind("<Double-1>", self.load_selected_product_add)
        self.refresh_product_list()

    def setup_withdraw_product_frame(self):
        self.frame_logo_withdraw = ctk.CTkFrame(self.frame_withdraw_product)
        self.frame_logo_withdraw.pack(pady=10, padx=10, fill="x")

        self.company_name_withdraw = ctk.CTkLabel(self.frame_logo_withdraw, text="Super Petrochemical Limited", font=("Arial", 24, "bold"))
        self.company_name_withdraw.pack(pady=5)
        self.department_name_withdraw = ctk.CTkLabel(self.frame_logo_withdraw, text="Instrument and Control Department", font=("Arial", 15, "bold"))
        self.department_name_withdraw.pack(pady=5)

        self.frame_top_withdraw = ctk.CTkFrame(self.frame_withdraw_product)
        self.frame_top_withdraw.pack(pady=10, padx=10, fill="x")
        
        self.frame_middle_withdraw = ctk.CTkFrame(self.frame_withdraw_product)
        self.frame_middle_withdraw.pack(pady=10, padx=10, fill="x")
        
        self.frame_bottom_withdraw = ctk.CTkFrame(self.frame_withdraw_product)
        self.frame_bottom_withdraw.pack(pady=10, padx=10, fill="both", expand=True)

        # Top Frame Widgets
        self.frame_entries_withdraw = ctk.CTkFrame(self.frame_top_withdraw)
        self.frame_entries_withdraw.pack(pady=10, padx=10)

        ctk.CTkLabel(self.frame_entries_withdraw, text="SR By:").grid(row=0, column=0, padx=5, pady=5)
        self.sr_by_requisition = ctk.CTkEntry(self.frame_entries_withdraw)
        self.sr_by_requisition.grid(row=0, column=1, padx=5, pady=5)

        ctk.CTkLabel(self.frame_entries_withdraw, text="Required for:").grid(row=0, column=2, padx=5, pady=5)
        self.required_requisition = ctk.CTkEntry(self.frame_entries_withdraw)
        self.required_requisition.grid(row=0, column=3, padx=5, pady=5)
        # Requisited Items Table
        self.tree_requisition = ttk.Treeview(self.frame_top_withdraw, columns=("SL No.", "Item No","Product Name","Quantity","Size","Remarks"), show="headings")
        self.tree_requisition.column('SL No.',width=30)
        self.tree_requisition.column('Item No',width=100)
        self.tree_requisition.column('Product Name',width=400)
        self.tree_requisition.column('Quantity',width=100)
        self.tree_requisition.column('Size',width=100)
        self.tree_requisition.column('Remarks',width=100)

        self.tree_requisition.heading("SL No.", text="Sl No.")
        self.tree_requisition.heading("Item No", text="Item No")
        self.tree_requisition.heading("Product Name", text="Product Name")
        self.tree_requisition.heading("Quantity", text="Quantity")
        self.tree_requisition.heading("Size", text="Size")
        self.tree_requisition.heading("Remarks", text="Remarks")
        self.tree_requisition.pack(fill="x", expand=True)
        self.tree_requisition.bind("<Delete>", lambda event: self.delete_selected())

        
        # ctk.CTkLabel(self.frame_entries_withdraw, text="Quantity:").grid(row=1, column=0, padx=5, pady=5)
        # self.entry_quantity_withdraw = ctk.CTkEntry(self.frame_entries_withdraw)
        # self.entry_quantity_withdraw.grid(row=1, column=1, padx=5, pady=5)

        self.button_withdraw_product = ctk.CTkButton(self.frame_top_withdraw, text="Withdraw Product", command=self.withdraw_popup)
        self.button_withdraw_product.pack( padx=10, pady=10)

        # Middle Frame Widgets
        ctk.CTkLabel(self.frame_middle_withdraw, text="Search Product:").grid(row=0, column=0, padx=5, pady=5)
        self.entry_search_withdraw = ctk.CTkEntry(self.frame_middle_withdraw)
        self.entry_search_withdraw.grid(row=0, column=1, padx=5, pady=5)
        
        ctk.CTkButton(self.frame_middle_withdraw, text="Search", command=self.search_product_withdraw).grid(row=0, column=2, padx=5, pady=5)
        ctk.CTkButton(self.frame_middle_withdraw, text="Clear Search", command=self.refresh_product_list_withdraw).grid(row=0, column=3, padx=5, pady=5)
        ctk.CTkLabel(self.frame_bottom_withdraw, text="Stock Items").pack(padx=10, pady = 10, fill="x")
        self.tree_withdraw = ttk.Treeview(self.frame_bottom_withdraw, columns=(
            "Item No", "Name","Quantity","Size","Type","Range","Store","Rack","Column","Row","Box", "Description", "Plant Name", "Part No", "Tag"), show="headings")
        
        verscrlbar = ttk.Scrollbar(self.frame_bottom_withdraw, 
                           orient ="vertical", 
                           command = self.tree_withdraw.yview)
        # Calling pack method w.r.to vertical 
        # scrollbar
        verscrlbar.pack(side ='right', fill ='both')
        #resizing the scrollbar according to table size
        self.tree_withdraw.configure(yscrollcommand=verscrlbar.set)

        # self.tree_withdraw.column('Item No',width=60)
        # self.tree_withdraw.column('Name',width=200)
        # self.tree_withdraw.column('Quantity',width=60)
        # self.tree_withdraw.column('Size',width=30)
        # self.tree_withdraw.column('Range',width=50)
        # self.tree_withdraw.column('Type',width=50)
        # self.tree_withdraw.column('Store',width=50)
        # self.tree_withdraw.column('Rack',width=50)
        # self.tree_withdraw.column('Column',width=50)
        # self.tree_withdraw.column('Box',width=50)
        # self.tree_withdraw.column('Description',width=250)
        

        self.tree_withdraw.heading("Item No", text="Item No")
        self.tree_withdraw.heading("Name", text="Product Name")
        self.tree_withdraw.heading("Quantity", text="Quantity")
        self.tree_withdraw.heading("Size", text="Size")
        self.tree_withdraw.heading("Type", text="Type")
        self.tree_withdraw.heading("Range", text="Range")
        self.tree_withdraw.heading("Store", text="Store")
        self.tree_withdraw.heading("Rack", text="Rack")
        self.tree_withdraw.heading("Column", text="Column")
        self.tree_withdraw.heading("Row", text="Row")
        self.tree_withdraw.heading("Box", text="Box")
        self.tree_withdraw.heading("Description", text="Description")
        self.tree_withdraw.heading("Plant Name", text="Plant Name")
        self.tree_withdraw.heading("Part No", text="Part No")
        self.tree_withdraw.heading("Tag", text="Tag")
        self.tree_withdraw.pack(fill="both", expand=True)

        # self.tree_withdraw.insert("", "end",("Item No","1"))
        ctk.CTkButton(self.frame_withdraw_product, text="Back to Home", command=self.show_home_frame).pack(pady=10)
        # ctk.CTkButton(self.frame_withdraw_product, text="Print", command=self.print_popup).pack(pady=10)

        self.tree_withdraw.bind("<Double-1>", lambda event: self.popup_quantity())

        # lambda event: self.popup_quantity
        # self.add_to_cart()
        self.refresh_product_list_withdraw()
        
    def delete_selected(self):
        print("Delete Pressed")
        pass
    def show_home_frame(self):
        self.frame_home.pack(pady=10, padx=10, fill="both", expand=True)
        self.frame_add_product.pack_forget()
        self.frame_withdraw_product.pack_forget()

    def show_add_product_frame(self):
        self.frame_add_product.pack(pady=10, padx=10, fill="both", expand=True)
        self.frame_home.pack_forget()
        self.frame_withdraw_product.pack_forget()
        self.refresh_product_list()

    def show_withdraw_product_frame(self):
        self.frame_withdraw_product.pack(pady=10, padx=10, fill="both", expand=True)
        self.refresh_product_list_withdraw()
        self.frame_home.pack_forget()
        self.frame_add_product.pack_forget()
    def export_to_excel(self):
        file_path = "products.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Products"
        headers = ["Item No", "Product Name", "Brand", "Model", "Quantity", "Size", "Type", "Range", "Store", "Rack", "Column", "Row", "Box", "Description", "Plant Name", "Part No", "Tag"]
        sheet.append(headers)

        for product in self.product_list:
            row = [product[header] for header in headers]
            sheet.append(row)

        try:
            workbook.save(file_path)
            # messagebox.showinfo("Export Successful", f"Data exported successfully to {file_path}")
        except Exception as e:
            messagebox.showerror("Export Error", f"An error occurred while exporting: {e}")
        
        workbook.save("products.xlsx")
        # messagebox.showinfo("Excel File", "Products saved to products.xlsx")

    def add_product(self):
        ### When product is loaded from the table, selected product will be updated
        if self.product_loaded_add == 1: 
            selected_item = self.tree_add.selection()
            if not selected_item:
                messagebox.showerror("Selection Error", "No item selected to update")
                return
            
            # print(f"Selected Item: {selected_item}")
            # print(f"Focused Item: {self.tree_add.focus()}")
            item_index_no = self.tree_add.focus()
            # print(f"Selected Product: {self.tree_add.item(item_index_no)}")
            loaded_product_item_no = self.tree_add.item(item_index_no)["values"][0]
            loaded_product_item_no = f"{loaded_product_item_no:06d}"
            # print(f"Loaded Product Item No: {loaded_product_item_no}")
            
            for index, product in enumerate(self.product_list):

                if (
                    str(loaded_product_item_no) in str(product["Item No"]).lower() ):
                    # print(f"Loaded Product::::::: {self.product_list[index]}")
                    selected_product = self.product_list[index]
                    item_index = index

            # print(f"Item Index No: {item_index}")
            product_name = self.entry_product_name_add.get()
            brand = self.entry_brand_add.get()
            model = self.entry_model_add.get()
            quantity = self.entry_quantity_add.get()
            size = self.entry_size_add.get()
            product_type = self.entry_type_add.get()
            product_range = self.entry_range_add.get()
            store = self.entry_store_add.get()
            rack = self.entry_rack_add.get()
            row = self.entry_row_add.get()
            column = self.entry_column_add.get()
            box_no = self.entry_box_no_add.get()
            description = self.entry_description_add.get("1.0", tk.END).strip()
            plant_name = self.entry_plant_name_add.get()
            tag = self.entry_tag_add.get()
            part_no = self.entry_part_no_add.get()
            if not part_no:
                part_no = self.generate_part_no(plant_name, product_type, product_range,brand,model)

            self.product_list[item_index] = {
                "Item No": loaded_product_item_no,
                "Product Name": product_name,
                "Brand": brand,
                "Model": model,
                "Quantity": quantity,
                "Size": size,
                "Type": product_type,
                "Range": product_range,
                "Store": store,
                "Rack": rack,
                "Row": row,
                "Column": column,
                "Box": box_no,
                "Description": description,
                "Plant Name": plant_name,
                "Part No": part_no,
                "Tag": tag
                }
            updated_list = self.product_list[int(loaded_product_item_no)-1]
            self.update_db('instrument_store.stocks', updated_list)
            # self.export_to_excel()
            self.refresh_product_list()
            self.clear_add_product_entries()
            

        else:
            product_name = self.entry_product_name_add.get()
            brand = self.entry_brand_add.get()
            model = self.entry_model_add.get()
            quantity = self.entry_quantity_add.get()
            size = self.entry_size_add.get()
            product_type = self.entry_type_add.get()
            product_range = self.entry_range_add.get()
            store = self.entry_store_add.get()
            rack = self.entry_rack_add.get()
            row = self.entry_row_add.get()
            column = self.entry_column_add.get()
            box_no = self.entry_box_no_add.get()
            description = self.entry_description_add.get("1.0", tk.END).strip()
            plant_name = self.entry_plant_name_add.get()
            tag = self.entry_tag_add.get()
            part_no = self.entry_part_no_add.get()
            if not part_no:
                part_no = self.generate_part_no(plant_name, product_type, product_range,brand,model)

            if not all([product_name, plant_name, store, product_type, size, product_range, quantity, rack, column, box_no, part_no]):
                messagebox.showerror("Error", "Please fill all the fields.")
                return
            
            item_no = f"{self.item_no_counter:06d}"
            # print(f"Counter No:: {self.item_no_counter}")
            self.product_list.append({
                "Item No": item_no,
                "Product Name": product_name,
                "Brand": brand,
                "Model": model,
                "Quantity": quantity,
                "Size": size,
                "Type": product_type,
                "Range": product_range,
                "Store": store,
                "Rack": rack,
                "Row": row,
                "Column": column,
                "Box": box_no,
                "Description": description,
                "Plant Name": plant_name,
                "Part No": part_no,
                "Tag": tag
            })

            self.add_to_db('instrument_store.stocks',self.product_list[self.item_no_counter-1])
            # self.export_to_excel()
            self.refresh_product_list()
            # self.item_no_counter += 1
            messagebox.showinfo("Success", "Product added successfully.")
            self.clear_add_product_entries()
    # Function to convert camel case to snake case
    def to_snake_case(self,s):
        s = s.lower()
        s = s.replace(' ', '_')
        s = s.replace('plant_name','plant')
        return s



    def create_db_connection(self, host_name, user_name, user_password, db_name):
        connection = None
        try:
            self.mydb = mysql.connector.connect(
                host=host_name,
                user=user_name,
                passwd=user_password,
                database=db_name
            )
            print("MySQL Database connection successful")
        except Error as err:
            print(f"Error: '{err}'")
        return self.mydb

    # def update_db(self,table,product):
    #     _product = product
    #     fields = list(_product.keys())
    #     values = list(_product.values())
    #     fields.pop(0)
    #     item_no = int(values.pop(0))
    #     print(item_no)
    #     print(f"Values:: {values[2]}")

    #     sqlFormula = f"""UPDATE instrument_store.stocks SET
    #         product_name = {values[0]},
    #         brand = '{values[1]}',
    #         model = '{values[2]}',
    #         quantity = '{values[3]}',
    #         size = '{values[4]}',
    #         type = '{values[5]}',
    #         range = '{values[6]}',
    #         store = '{values[7]}',
    #         rack = '{values[8]}',
    #         column = '{values[9]}',
    #         row = '{values[10]}',
    #         box = '{values[11]}',
    #         description = '{values[12]}',
    #         plant = '{values[13]}',
    #         part_no = '{values[14]}',
    #         tag = '{values[15]}',
    #         WHERE item_no = {item_no}

    #     """

    #     print(f"SQL Query:: {sqlFormula}")
    #     mycursor = self.mydb.cursor()        
    #     mycursor.execute(sqlFormula)
    #     self.mydb.commit()
    def update_db(self, table, product):
        print(product)
        _product = product
        fields = list(_product.keys())
        values = list(_product.values())
        fields.pop(0)
        item_no = int(values.pop(0))
        print(item_no)
        print(f"Values:: {values[2]}")

        for field, value in zip(self.db_fields, values):
            print(f"{field} : {value}")
            mycursor = self.mydb.cursor()
            sqlFormula = f"UPDATE {table} SET {field}= %s WHERE item_no = %s"
            try:
                mycursor.execute(sqlFormula, (value, item_no))
                self.mydb.commit()
                print(f"Updated {field} = {value} for item_no = {item_no}")
            except mysql.connector.Error as err:
                print(f"Error: {err}")
                self.mydb.rollback()
        
        mycursor.close()
        # Ensure the product dictionary is not empty
        # if not _product:
        #     raise ValueError("Product dictionary must contain values to update")

        

    def add_to_db(self,table,product):
        values = tuple(product.values())
        print(f"value to added:: {values}") 
        mycursor = self.mydb.cursor()
        formula =f"INSERT INTO {table} VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
        mycursor.execute(formula,values)
        self.mydb.commit()

        # mycursor.execute(f""/INSERT INTO 'stocks' {self.db_fields} VALUES (3,'PT', 'YOKOGAWA', 'EJ', 2, '0.5', '-', '24V', '1', '2', '3', '4', '6', 'XXX', '3', '3')/""")
        # mycursor.execute("INSERT INTO `instrument_store`.`stocks` (`item_no`, `product_name`, `brand`, `model`, `quantity`, `size`, `type`, `range`, `store`, `rack`, `row`, `column`, `box`, `description`, `plant`, `tag`) VALUES ('3', 'FT', 'yokogawa', 'EJX', '3', '0.5', 'v', '24V', '1', '5', '6', '3', '1', '4', '5', '3')")



    # def execute_query(self, query, data):
    #     cursor = self.db_connection.cursor()
    #     try:
    #         cursor.execute(query, data)
    #         self.db_connection.commit()
    #         print("Query successful")
    #     except Error as err:
    #         print(f"Error: '{err}'")


    def load_from_excel(self):
        pass
        # self.load_from_db('instrument_store.stocks')
        # file_path = "products.xlsx"
        # if not file_path:
        #     return
        
        # try:
        #     workbook = openpyxl.load_workbook(file_path)
        #     sheet = workbook.active

        #     self.product_list = []
        #     self.item_no_counter = 1

        #     for row in sheet.iter_rows(min_row=2, values_only=True):
        #         product = {
        #             "Item No": row[0],
        #             "Product Name": row[1],
        #             "Brand": row[2],
        #             "Model": row[3],
        #             "Quantity": row[4],
        #             "Size": row[5],
        #             "Type": row[6],
        #             "Range": row[7],
        #             "Store": row[8],
        #             "Rack": row[9],
        #             "Column": row[10],
        #             "Row": row[11],
        #             "Box": row[12],
        #             "Description": row[13],
        #             "Plant Name": row[14],
        #             "Part No": row[15],
        #             "Tag": row[16]
        #         }
        #         self.product_list.append(product)
        #         self.item_no_counter += 1
        #     self.replace_none_with_empty_string(self.product_list)
        #     self.refresh_product_list()
        #     self.refresh_product_list_withdraw()
        #     # messagebox.showinfo("Load Successful", f"Data loaded successfully from {file_path}")
        # except Exception as e:
        #     messagebox.showerror("Load Error", f"An error occurred while loading: {e}")
    def load_from_db(self,table):
        connection = self.mydb.cursor()
        sql_formula = f"SELECT * FROM {table}"
        connection.execute(sql_formula)
        records = connection.fetchall()
        # print(f"Records:: {records}")

        for items in records:
            products = {
            "Item No"       :   f"{items[0]:06d}",
            "Product Name"  :items[1],   
            "Brand"     :items[2],
            "Model"     :items[3], 
            "Quantity"  :items[4],
            "Size"      :items[5], 
            "Type"      :items[6], 
            "Range"     :items[7], 
            "Store"     :items[8], 
            "Rack"      :items[9], 
            "Column"    :items[10], 
            "Row"       :items[11], 
            "Box"       :items[12], 
            "Description":items[13], 
            "Plant Name" :items[14], 
            "Part No"    :items[15], 
            "Tag"       :items[16]
            }
            self.item_no_counter += 1
            self.product_list.append(products)
        # print(f"Product list--->> {self.product_list}")

    def replace_none_with_empty_string(self,product_list):
        for product in product_list:
            for key, value in product.items():
                if value is None:
                    product[key] = ""

    def search_product_withdraw(self):
            # search_term = self.entry_search_add.get() if isinstance(self.entry_search_add.get(), str) else str(self.entry_search_add.get()).lower()
            search_term = self.entry_search_withdraw.get() if isinstance(self.entry_search_withdraw.get(), str) else str(self.entry_search_withdraw.get()).lower()
            self.searched_withdraw=1
            for row in self.tree_add.get_children():
                self.tree_add.delete(row)
            for row in self.tree_withdraw.get_children():
                self.tree_withdraw.delete(row)
            
            for product in self.product_list:
                if (
                    search_term in str(product["Item No"]) or
                    search_term in product["Product Name"].lower() or
                    search_term in product["Plant Name"].lower() or
                    search_term in product["Type"].lower() or
                    search_term in product["Description"].lower() or
                    search_term in product["Tag"].lower() or
                    search_term in str(product["Size"]).lower() or
                    search_term in product["Part No"].lower()):
                    # refreshed_list.append(product)
                    # print(f"Refreshed List: {refreshed_list}")
                    self.tree_withdraw.insert("", "end", values=(
                        product["Item No"], 
                        product["Product Name"], 
                        product["Brand"], 
                        product["Model"], 
                        product["Quantity"],
                        product["Size"], 
                        product["Type"], 
                        product["Range"], 
                        product["Store"], 
                        product["Rack"], 
                        product["Column"], 
                        product["Row"], 
                        product["Box"], 
                        product["Description"], 
                        product["Plant Name"], 
                        product["Part No"], 
                        product["Tag"]
                    ))

                    
    def search_product_add(self):
            search_term = self.entry_search_add.get() if isinstance(self.entry_search_add.get(), str) else str(self.entry_search_add.get()).lower()
            # search_term = self.entry_search_withdraw.get() if isinstance(self.entry_search_withdraw.get(), str) else str(self.entry_search_withdraw.get()).lower()
            self.searched_add=1
            search_term = search_term.lower() # Convert all characters from search entry to lower-case
            for row in self.tree_add.get_children():
                self.tree_add.delete(row)
            for row in self.tree_withdraw.get_children():
                self.tree_withdraw.delete(row)
            
            for product in self.product_list:
                if (
                    search_term in str(product["Item No"]) or
                    search_term in product["Product Name"].lower() or
                    search_term in product["Brand"].lower() or
                    search_term in product["Model"].lower() or
                    search_term in product["Plant Name"].lower() or
                    search_term in product["Type"].lower() or
                    search_term in product["Description"].lower() or
                    search_term in product["Tag"].lower() or
                    search_term in str(product["Size"]).lower() or
                    search_term in product["Part No"].lower()):
                    # refreshed_list.append(product)
                    # print(f"Refreshed List: {refreshed_list}")
                    self.tree_add.insert("", "end", values=(
                        product["Item No"], 
                        product["Product Name"], 
                        product["Brand"], 
                        product["Model"], 
                        product["Quantity"],
                        product["Size"], 
                        product["Type"], 
                        product["Range"], 
                        product["Store"], 
                        product["Rack"], 
                        product["Column"], 
                        product["Row"], 
                        product["Box"], 
                        product["Description"], 
                        product["Plant Name"], 
                        product["Part No"], 
                        product["Tag"]
                    ))
            
    def refresh_product_list(self):
        self.entry_search_add.delete(0, 'end')
        self.product_loaded_add = 0
        self.button_add_product.configure(text ="Add Product")
        for i in self.tree_add.get_children():
            self.tree_add.delete(i)
        for products in self.product_list:
            self.tree_add.insert("", "end", values=(
            products["Item No"], 
            products["Product Name"], 
            products["Brand"], 
            products["Model"], 
            products["Quantity"], 
            products["Size"], 
            products["Type"], 
            products["Range"], 
            products["Store"], 
            products["Rack"], 
            products["Column"], 
            products["Row"], 
            products["Box"], 
            products["Description"], 
            products["Plant Name"], 
            products["Part No"], 
            products["Tag"]
        ))

    def refresh_product_list_withdraw(self):
        self.searched_withdraw =0
        # self.load_from_excel()                                                                                                        
        for i in self.tree_withdraw.get_children():
            self.tree_withdraw.delete(i)
        # print(f"Product List:{self.product_list}")
        for products in self.product_list:
            self.tree_withdraw.insert("", "end", values=(
                    products["Item No"], products["Product Name"], products["Quantity"],products["Size"], products["Type"], products["Range"], products["Store"], products["Rack"], products["Column"],products["Row"], products["Box"], products["Description"], products["Plant Name"], products["Part No"], products["Tag"]
            ))

       
    def load_selected_product_add(self, event):
        
        selected_item = self.tree_add.selection()
        if not selected_item:
            return
        # print(f"Selected Item: {selected_item}")
        # print(f"Focused Item: {self.tree_add.focus()}")
        item_index_no = self.tree_add.focus()
        # print(f"Selected Product: {self.tree_add.item(item_index_no)}")
        loaded_product_item_no = self.tree_add.item(item_index_no)["values"][0]
        loaded_product_item_no = f"{loaded_product_item_no:06d}"
        # print(f"Loaded Product Item No: {loaded_product_item_no}")
        
        for index, product in enumerate(self.product_list):

            if (
                str(loaded_product_item_no) in str(product["Item No"]).lower() ):
                # print(f"Loaded Product::::::: {self.product_list[index]}")
                selected_product = self.product_list[index]

        self.product_loaded_add = 1
        self.button_add_product.configure(text="Update")

        self.entry_product_name_add.delete(0, tk.END)
        if selected_product["Product Name"] is None:
            self.entry_product_name_add.insert(0, "")
        else:

            self.entry_product_name_add.insert(0, selected_product["Product Name"])

        self.entry_brand_add.delete(0, tk.END)
        if selected_product["Brand"] is None:
            self.entry_brand_add.insert(0, "")
        else:
            self.entry_brand_add.insert(0, selected_product["Brand"])

        self.entry_model_add.delete(0, tk.END)
        if selected_product["Model"] is None:
            self.entry_model_add.insert(0, "")
        else:
            self.entry_model_add.insert(0, selected_product["Model"])

        self.entry_quantity_add.delete(0, tk.END)

        if selected_product["Quantity"] is None:
            self.entry_quantity_add.insert(0, "")
        else:
            self.entry_quantity_add.insert(0, selected_product["Quantity"])

        self.entry_type_add.set(selected_product["Type"])

        self.entry_range_add.delete(0, tk.END)

        if selected_product["Range"] is None:
            self.entry_range_add.insert(0, "")
        else:
            self.entry_range_add.insert(0, selected_product["Range"])

        self.entry_store_add.delete(0, tk.END)

        if selected_product["Store"] is None:
            self.entry_store_add.insert(0,"")
        else:
            self.entry_store_add.insert(0, selected_product["Store"])

        self.entry_rack_add.delete(0, tk.END)
        self.entry_rack_add.insert(0, selected_product["Rack"])

        self.entry_column_add.delete(0, tk.END)
        self.entry_column_add.insert(0, selected_product["Column"])

        self.entry_box_no_add.delete(0, tk.END)
        self.entry_box_no_add.insert(0, selected_product["Box"])

        self.entry_plant_name_add.set(selected_product["Plant Name"])

        self.entry_row_add.delete(0, tk.END)
        self.entry_row_add.insert(0, selected_product["Row"])

        self.entry_tag_add.delete(0, tk.END)

        if selected_product["Tag"] is None:
            self.entry_tag_add.insert(0,"")
        else:
            self.entry_tag_add.insert(0, selected_product["Tag"])

        self.entry_part_no_add.delete(0, tk.END)
        self.entry_part_no_add.insert(0, selected_product["Part No"])

        self.entry_size_add.delete(0, tk.END)
        self.entry_size_add.insert(0, selected_product["Size"])

        self.entry_description_add.delete('1.0', tk.END)
        if selected_product["Description"] is None:
            self.entry_description_add.insert("1.0", "")
        else:
            self.entry_description_add.insert("1.0", selected_product["Description"])


    def clear_add_product_entries(self):
        self.entry_product_name_add.delete(0, 'end')
        self.entry_brand_add.delete(0, 'end')
        self.entry_model_add.delete(0, 'end')
        self.entry_part_no_add.delete(0, 'end')
        self.entry_tag_add.delete(0, 'end')
        self.entry_plant_name_add.set("")
        self.entry_store_add.delete(0, 'end')
        self.entry_type_add.set("")
        self.entry_size_add.delete(0, 'end')
        self.entry_range_add.delete(0, 'end')
        self.entry_quantity_add.delete(0, 'end')
        self.entry_rack_add.delete(0, 'end')
        self.entry_column_add.delete(0, 'end')
        self.entry_row_add.delete(0, 'end')
        self.entry_description_add.delete('1.0', 'end')
        self.entry_box_no_add.delete(0, 'end')
        self.refresh_product_list()
    def popup_quantity(self):
        popup = tk.Toplevel(self.root)
        popup.title("Add Quantity")
        popup.geometry("300x200")
        print("Initiated")

        def add_to_cart_():
            self.add_to_cart(quantity.get())
            # Code to handle printing can go here
            print("Added To Cart")
            popup.destroy()
        def cancel_popup():
            print("Cancelled")
            popup.destroy()

        # Add print options here (e.g., selection of specific details, number of copies)
        tk.Label(popup, text="Quantity").pack(pady=5)
        quantity = ctk.CTkEntry(popup)
        quantity.pack(pady=5)
        tk.Button(popup, text="Add to Cart", command=add_to_cart_).pack(pady=10)
        tk.Button(popup, text="Cancel", command=cancel_popup).pack(pady=10)
        
        

    def generate_part_no(self, plant_name, type_, range_,_brand,_model):
    # Function to remove vowels from a string
        def remove_vowels(s):
            vowels = "AEIOUaeiou"
            return "".join([char for char in s if char not in vowels])

        # Mapping of plant codes
        plant_mapping = {
            "CHEMEX": "CHE",
            "ZEHUA": "ZEH",
            "BTX": "BTX",
            "HEXANE": "HEX",
            "LCP": "LCP",
            "TNS": "TNS",
            "REFORMER 2": "REF02",
            "CFU-10000BPD": "CFU",
            "FLARE": "FLR",
            "K601A": "K601A",
            "K601B": "K601B",
            "K401A": "K401A",
            "K401B": "K401B",
            "K401C": "K401C",
            "COOLING TOWER" :"C-TOWER",
            "RO" : "RO"
        }

        # Get plant code from mapping or default to first three letters uppercase without vowels
        plant_code = plant_mapping.get(plant_name.upper(), remove_vowels(plant_name.upper())[:5])

        # Remove vowels from plant_name and type_
        # plant_name_without_vowels = remove_vowels(plant_name.upper())
        # type_without_vowels = remove_vowels(type_.upper())

        # Construct part number based on whether brand is provided
        if _brand and _model:
            part_no = f"{plant_code}-{type_}-{_brand}-{_model}-{range_}"
        else:
            part_no = f"{plant_code}-{type_}-{range_}"
    # Construct part number in desired format
    
        return part_no
        
    def add_to_cart(self,added_quantity):
        if not self.cart_list:
            self.previous_product_list = self.product_list.copy()  # Save a copy of the current product list
        
        selected_product = dict()
        item_index=0
        i=0
        selected_item = self.tree_withdraw.selection()
        if not selected_item:
            return
        if self.searched_withdraw ==1 :
                selected_item_index = self.tree_withdraw.focus()
                curntItemPartNo = self.tree_withdraw.item(selected_item_index)["values"][0]
                curntItemPartNo = f"{curntItemPartNo:06d}"
                print(f"Selected Item: {curntItemPartNo}")

                for index, product in enumerate(self.product_list):
                    if (
                        str(curntItemPartNo) in str(product["Item No"]).lower() ):
                        print(f"Loaded Product::::::: {self.product_list[index]}")
                        selected_product = self.product_list[index]
                        item_index = index
                        selected_product = self.product_list[item_index]          
                    
        if self.searched_withdraw ==0 :
                selected_item_index = self.tree_withdraw.focus()
                curntItemPartNo = self.tree_withdraw.item(selected_item_index)["values"][0]
                curntItemPartNo = f"{curntItemPartNo:06d}"
                print(f"Selected Item: {curntItemPartNo}")
                for index, product in enumerate(self.product_list):
                    if (
                        str(curntItemPartNo) in str(product["Item No"]).lower() ):
                        print(f"Loaded Product::::::: {self.product_list[index]}")
                        selected_product = self.product_list[index]
                        item_index = index
                        selected_product = self.product_list[item_index]

                # curntItem = self.tree_withdraw.focus()
                # # print(f"Current Item {curntItem}")
                # item_index = self.tree_withdraw.index(curntItem)
                # curntItem = self.tree_withdraw.item(curntItem)["values"]      
                # selected_product = self.product_list[item_index]


        self.searched_withdraw ==0
        
        if(self.tree_requisition.get_children() == "NULL"):
            self.requisition_item_counter=1

        stock_quantity = selected_product["Quantity"]
        remaining = int(stock_quantity) - int(added_quantity)
        print(f"Stock Quantity: {stock_quantity}")
        print(f"Remaining Quantity: {remaining}")
        if (remaining) < 0:
            print("Low Stock")
            messagebox.showwarning("Warning!","Low Stock. Please try again")
            return
        self.product_list[item_index]["Quantity"] = str(int(self.product_list[item_index]["Quantity"]) - int(added_quantity))
        
        cart = {
            "SL No." :  self.requisition_item_counter,
            "Item No" : selected_product["Item No"],
            "Product Name": selected_product["Product Name"],
            "Quantity" : added_quantity,
            "Size" : selected_product["Size"]
        }
        self.cart_list.append(cart)
        # print(self.cart_list)
        # print(f"My Cart: {self.cart_list}")

        self.requisition_item_counter +=1
        self.export_to_excel()
        self.refresh_cart()
        self.refresh_product_list_withdraw()

    def refresh_cart(self):
        for i in self.tree_requisition.get_children():
            self.tree_requisition.delete(i)
        for product in self.cart_list:
            self.tree_requisition.insert("", "end", values=(
                product["SL No."], product["Item No"], product["Product Name"], product["Quantity"],product["Size"]))

    def find_missing_number(self):
        available_numbers = [325645, 365962, 259954, 256565, 201251, 689845, 875459, 565656]
        for product in self.product_list:
            self.numbers.append(product["Item No"])
        print(self.numbers)
        max_number = max(self.numbers)
        print(f"Max Number is : {max_number}")
        # Pop elements from available_numbers and remove them from numbers
        while available_numbers:
            available_number = available_numbers.pop(0)  # Pop the first element from available_numbers
            if available_number in self.numbers:
                self.numbers.remove(available_number)  # Remove the element from numbers

        print("Updated numbers list:", self.numbers)
        print("Updated available numbers list:", available_numbers)


    def print_popup(self):
        popup = tk.Toplevel(self.root)
        popup.title("Print")
        popup.geometry("300x200")

        def print_details():
            # Code to handle printing can go here
            print("Print initiated")
            popup.destroy()

        tk.Label(popup, text="Select print options:", font=("Arial", 14)).pack(pady=10)

        # Add print options here (e.g., selection of specific details, number of copies)
        tk.Label(popup, text="Number of copies:").pack(pady=5)
        num_copies = tk.Entry(popup)
        num_copies.pack(pady=5)

        tk.Button(popup, text="Print", command=print_details).pack(pady=10)
        tk.Button(popup, text="Cancel", command=popup.destroy).pack(pady=10)
    def withdraw_popup(self):
        popup = tk.Toplevel(self.root)
        popup.title("Print")
        popup.geometry("300x200")

        def confirm_withdraw():
            # Code to handle printing can go here
            print("Withdraw Confirmed")
            popup.destroy()
            self.refresh_product_list_withdraw()
        def cancel_withdraw():
            # Code to handle printing can go here
            self.product_list = self.previous_product_list.copy()
            print("Withdraw Cancelled")
            # print(f"Current Product List::: {self.product_list}")
            # print(f"Previous product list ::::: {self.previous_product_list}")
            popup.destroy()     
            self.refresh_product_list_withdraw()

        # tk.Label(popup, text="Select print options:", font=("Arial", 14)).pack(pady=10)

        # # Add print options here (e.g., selection of specific details, number of copies)
        # tk.Label(popup, text="Number of copies:").pack(pady=5)
        num_copies = tk.Entry(popup)
        num_copies.pack(pady=5)

        tk.Button(popup, text="Confirm", command=confirm_withdraw).pack(pady=10)
        tk.Button(popup, text="Cancel", command=cancel_withdraw).pack(pady=10)
    
        
if __name__ == "__main__":
    ctk.set_appearance_mode("Light")  # Modes: "System" (default), "Dark", "Light"
    ctk.set_default_color_theme("green")  # Themes: "blue" (default), "green", "dark-blue"
    root = ctk.CTk()
    app = StoreManagementApp(root)
    root.mainloop()